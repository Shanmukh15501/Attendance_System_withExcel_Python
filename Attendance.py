import cv2
import numpy as np
import xlrd
from openpyxl import load_workbook
import time
import datetime

face_cascade=cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('MyTraining.xml')
Video = cv2.VideoCapture(0)
print("Do you want to change Sheet or Update 1 to Change 0 to Update")
op = int (input("Enter a value"))
op1 = xlrd.open_workbook('Output.xlsx')
wb2 = load_workbook('Output.xlsx')

if (op ==  1):
      a=datetime.date.today()
      b=a
      print(b)
      wb2.create_sheet(str(b))
      sheet = wb2[str(b)]
      c1 = sheet.cell(row = 1, column = 1) 
      c1.value = "Name"
      c2 = sheet.cell(row= 1 , column = 2) 
      c2.value = "Date & Time"
else:
      hell = op1.sheet_names()
      hello = len(hell)-1
      sheet = wb2[hell[hello]]
      print(sheet)
      print(op1.sheet_names())
      c1 = sheet.cell(row = 1, column = 1) 
      c1.value = "Name"
      c2 = sheet.cell(row= 1 , column = 2) 
      c2.value = "Date & Time"
      a='A'
      b='B'
      row_count = sheet.max_row
      print("hello",row_count)
     
def update_Sheet(text):
      a='A'
      b='B'
      timedt=datetime.datetime.now()
      row_count = sheet.max_row
      print('row_count',row_count)
      row_count = (row_count)+1
      a=a+str(row_count)
      b=b+str(row_count)
      sheet[a]=text
      sheet[b]=str(timedt)
      wb2.save('Output.xlsx')

def output(id,text,x,y):
     cv2.putText(frame, text,(x, y-4), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,127,255), 1, cv2.LINE_AA)
      


def face(faces):
    dimension=len(faces)
    if (dimension!=0):
        for (x,y,w,h) in faces:
            cv2.rectangle(grey_img,(x,y),(x+w,y+h),(0,255,0),2)
            #roi_color=frame[y:y+h,x:x+w]
            id,confidence = recognizer.predict(grey_img[y:y+h,x:x+w])
            #print('id is',id)
            #print("confidance is",confidence)
            if id == 1:
                 text='Shanmukh'
                 output(id,text,x,y)
                 key=cv2.waitKey(2)
                 if key == ord('k'):
                      keynote = input('Enter K to enter into Excel')
                      if keynote == 'k':
                           print("Data Entered Successfully",text)
                           update_Sheet(text)
                           wb2.save('Output.xlsx')
            elif id == 2:
                 text='Obama'
                 output(id,text,x,y)
                 key=cv2.waitKey(2)
                 if key == ord('k'):
                      keynote = input('Enter K to enter into Excel')
                      if keynote == 'k':
                           print("Data Entered Successfully",text)
                           update_Sheet(text)
                           wb2.save('Output.xlsx'
            else:
                 text="unknown"
                 output(id,text,x,y)
                 key=cv2.waitKey(2)
                 if key == ord('k'):
                      keynote = input('Enter K to enter into Excel')
                      if keynote == 'k':
                           print("Data Entered Successfully",text)
                           update_Sheet(text)
                           wb2.save('Output.xlsx'
while True:
    check,frame=Video.read()
    grey_img=cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
    faces=face_cascade.detectMultiScale(frame)
    face(faces)
    cv2.imshow('Captures',frame)
    key=cv2.waitKey(2)
    if key == ord('q'):
        
        break    
Video.release()
cv2.destroyAllWindows()
